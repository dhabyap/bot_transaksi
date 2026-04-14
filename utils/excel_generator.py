import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def generate_excel_report(transactions, user_label, period_label, include_user_info=False):
    """
    Unified Excel generator for transaction reports.
    
    Args:
        transactions (list): List of transaction dictionaries.
        user_label (str): Name/Label of the user.
        period_label (str): Description of the period (e.g., 'April 2024').
        include_user_info (bool): Whether to include 'User' and 'Username' columns (Dashboard admin view).
        
    Returns:
        openpyxl.Workbook: The generated workbook object.
    """
    wb = openpyxl.Workbook()

    # ── Warna & Style ──────────────────────────────────────────────────
    header_fill  = PatternFill("solid", fgColor="1C1F2E")
    pemasukan_fill   = PatternFill("solid", fgColor="D1FAE5")   # hijau muda
    pengeluaran_fill = PatternFill("solid", fgColor="FEE2E2")   # merah muda
    investasi_fill   = PatternFill("solid", fgColor="DBEAFE")   # biru muda
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font():
        return Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    def body_font():
        return Font(name="Calibri", size=10)

    # ── Sheet 1: Transaksi ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Transaksi"

    # Baris judul
    # Hitung jumlah kolom: No(1), ID(2), Waktu(3), [User(4), Username(5) if True], Tipe(4/6), Item(5/7), Kategori(6/8), Nominal(7/9)
    col_count = 9 if include_user_info else 7
    title_range = f"A1:{get_column_letter(col_count)}1"
    gen_range = f"A2:{get_column_letter(col_count)}2"
    
    ws.merge_cells(title_range)
    title_cell = ws["A1"]
    title_cell.value = f"Laporan Keuangan — {user_label} | {period_label}"
    title_cell.font  = Font(name="Calibri", bold=True, size=13, color="1C1F2E")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(gen_range)
    gen_cell = ws["A2"]
    # Tambahkan deteksi platform
    platform = "Web Dashboard" if include_user_info else "Telegram Bot"
    gen_cell.value = f"Dibuat pada: {datetime.now().strftime('%d %B %Y, %H:%M')} via {platform}"
    gen_cell.font  = Font(name="Calibri", italic=True, size=9, color="888888")
    gen_cell.alignment = Alignment(horizontal="center")

    # Header kolom
    if include_user_info:
        headers = ["No", "ID Transaksi", "Tanggal & Waktu", "User", "Username", "Tipe", "Item / Keterangan", "Kategori", "Nominal (Rp)"]
        col_widths = [6, 14, 20, 20, 18, 15, 35, 20, 20]
    else:
        headers = ["No", "ID", "Tanggal & Waktu", "Tipe", "Item / Keterangan", "Kategori", "Nominal (Rp)"]
        col_widths = [5, 10, 20, 15, 35, 20, 20]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.fill      = header_fill
        cell.font      = hdr_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[4].height = 22

    # Isi data
    tipe_fill_map = {
        "pemasukan":   pemasukan_fill,
        "pengeluaran": pengeluaran_fill,
        "investasi":   investasi_fill,
    }

    for row_no, tx in enumerate(transactions, 1):
        r = row_no + 4
        tipe_str = (tx.get("tipe") or "").lower()
        row_fill = tipe_fill_map.get(tipe_str, None)

        ts = tx.get("timestamp")
        ts_str = ts.strftime("%d/%m/%Y %H:%M") if ts else "-"
        nominal = tx.get("nominal") or 0

        if include_user_info:
            nama = tx.get("first_name") or ""
            if tx.get("last_name"):
                nama += " " + tx["last_name"]
            row_data = [
                row_no,
                f"T-{tx.get('id', '')}",
                ts_str,
                nama.strip() or "-",
                f"@{tx.get('username')}" if tx.get("username") else "-",
                tipe_str.capitalize() or "-",
                tx.get("item") or "-",
                (tx.get("kategori") or "-").replace("_", " ").capitalize(),
                nominal,
            ]
        else:
            row_data = [
                row_no,
                f"T-{tx.get('id', '')}",
                ts_str,
                tipe_str.capitalize() or "-",
                tx.get("item") or "-",
                (tx.get("kategori") or "-").replace("_", " ").capitalize(),
                nominal,
            ]

        nominal_col = 9 if include_user_info else 7
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border = border
            cell.font   = body_font()
            if row_fill:
                cell.fill = row_fill
            
            if ci == nominal_col:   # Nominal
                cell.alignment     = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            elif ci == 1:           # No
                cell.alignment     = Alignment(horizontal="center")
            else:
                cell.alignment     = Alignment(horizontal="left", vertical="center")

    # Lebar kolom
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"

    # ── Sheet 2: Ringkasan ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Ringkasan")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 22

    summary_title = ws2.cell(row=1, column=1, value="Ringkasan Keuangan")
    summary_title.font = Font(name="Calibri", bold=True, size=13, color="1C1F2E")
    ws2.merge_cells("A1:B1")
    summary_title.alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 24

    ws2.cell(row=2, column=1, value="Periode").font = Font(bold=True, size=9, color="888888")
    ws2.cell(row=2, column=2, value=f"{user_label} | {period_label}").font = Font(size=9, color="888888")

    pemasukan   = sum(t.get("nominal", 0) or 0 for t in transactions if (t.get("tipe") or "").lower() == "pemasukan")
    pengeluaran = sum(t.get("nominal", 0) or 0 for t in transactions if (t.get("tipe") or "").lower() == "pengeluaran")
    investasi   = sum(t.get("nominal", 0) or 0 for t in transactions if (t.get("tipe") or "").lower() == "investasi")
    saldo_cash   = pemasukan - pengeluaran
    saldo_bersih = pemasukan - pengeluaran - investasi

    summary_rows = [
        ("Total Pemasukan",   pemasukan,   "3D9970"),
        ("Total Pengeluaran", pengeluaran, "E74C3C"),
        ("Total Investasi / Tabungan", investasi, "3498DB"),
        ("Saldo Cash (Pem. − Peng.)", saldo_cash, "2C3E50"),
        ("Saldo Bersih (inc. Investasi)", saldo_bersih, "8E44AD"),
    ]
    for ridx, (label, val, color) in enumerate(summary_rows, 4):
        lc = ws2.cell(row=ridx, column=1, value=label)
        lc.font   = Font(name="Calibri", bold=True, size=10, color=color)
        lc.border = border
        vc = ws2.cell(row=ridx, column=2, value=val)
        vc.font          = Font(name="Calibri", size=10)
        vc.border        = border
        vc.number_format = '#,##0'
        vc.alignment     = Alignment(horizontal="right")

    ws2.cell(row=10, column=1, value="Total Transaksi").font = Font(bold=True, size=10)
    wc = ws2.cell(row=10, column=2, value=len(transactions))
    wc.font      = Font(size=10)
    wc.alignment = Alignment(horizontal="center")

    return wb
