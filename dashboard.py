import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, Response
import database
from datetime import datetime

app = Flask(__name__)
app.secret_key = "super_secret_key_finance_bot_2024"

# Custom Jinja2 filter: parse JSON string di template
app.jinja_env.filters['from_json'] = json.loads

# Inisialisasi database MySQL saat aplikasi web mulai jalan
try:
    database.init_db()
except Exception as e:
    print(f"Gagal menginisialisasi database: {e}")


# ─── Helper ───────────────────────────────────────────────────────────────────

def _parse_filters(request):
    """Ambil month dan user_id filter dari query string."""
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    user_id_filter = request.args.get('user_id', '')
    uid = int(user_id_filter.strip()) if user_id_filter.strip().isdigit() else None
    return month, user_id_filter, uid


def _build_daily_chart_data(daily_trend, month_str):
    """
    Ubah hasil get_daily_trend menjadi dict berisi label + datasets
    yang siap di-serialize ke JSON untuk Chart.js.
    """
    # Tentukan jumlah hari dalam bulan
    year, month = map(int, month_str.split('-'))
    import calendar
    total_days = calendar.monthrange(year, month)[1]
    labels = list(range(1, total_days + 1))

    pemasukan_map  = {row['day']: float(row['total']) for row in daily_trend if row['tipe'] == 'pemasukan'}
    pengeluaran_map = {row['day']: float(row['total']) for row in daily_trend if row['tipe'] == 'pengeluaran'}
    investasi_map  = {row['day']: float(row['total']) for row in daily_trend if row['tipe'] == 'investasi'}

    return {
        'labels': labels,
        'pemasukan':   [pemasukan_map.get(d, 0)   for d in labels],
        'pengeluaran': [pengeluaran_map.get(d, 0) for d in labels],
        'investasi':   [investasi_map.get(d, 0)   for d in labels],
    }


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    month, user_id_filter, uid = _parse_filters(request)

    try:
        transactions  = database.get_transactions_by_month(month, uid)
        report        = database.get_monthly_report(uid, month)
        all_users     = database.get_all_users()
        category_data = database.get_category_breakdown(month, uid)
        daily_trend   = database.get_daily_trend(month, uid)
        stats         = database.get_stats_summary(month, uid)

        pemasukan   = float(report.get('pemasukan') or 0)
        pengeluaran = float(report.get('pengeluaran') or 0)
        investasi   = float(report.get('investasi') or 0)

    except Exception as e:
        print(f"Error accessing database in dashboard: {e}")
        flash("Terjadi kesalahan saat mengakses data dari database.", "error")
        transactions  = []
        all_users     = []
        category_data = []
        daily_trend   = []
        stats         = {'total_tx': 0, 'avg_daily_pengeluaran': 0, 'top_category': '-', 'top_category_val': 0}
        pemasukan = pengeluaran = investasi = 0.0

    # Kalkulasi saldo
    saldo_cash   = pemasukan - pengeluaran          # uang cash fisik
    saldo_bersih = pemasukan - pengeluaran - investasi  # termasuk investasi

    # Siapkan data chart sebagai JSON string agar aman di-inject ke JS
    pie_chart = {
        'labels': [row['kategori'] for row in category_data],
        'data':   [float(row['total']) for row in category_data],
    }
    daily_chart = _build_daily_chart_data(daily_trend, month)

    return render_template(
        'index.html',
        # Data transaksi
        transactions=transactions,
        all_users=all_users,
        # Filter state
        month=month,
        user_id_filter=user_id_filter,
        # Kartu ringkasan
        pemasukan=pemasukan,
        pengeluaran=pengeluaran,
        investasi=investasi,
        saldo_cash=saldo_cash,
        saldo_bersih=saldo_bersih,
        # Statistik
        stats=stats,
        # Chart data (JSON string)
        pie_chart_json=json.dumps(pie_chart),
        daily_chart_json=json.dumps(daily_chart),
    )


@app.route('/delete/<int:tx_id>', methods=['POST'])
def delete_tx(tx_id):
    """Dashboard adalah halaman admin — boleh hapus transaksi siapapun."""
    try:
        success = database.admin_delete_transaction(tx_id)
        if success:
            flash("✅ Data transaksi berhasil dihapus!", "success")
        else:
            flash("❌ Data transaksi tidak ditemukan.", "error")
    except Exception as e:
        print(f"Error in dashboard deletion: {e}")
        flash("Terjadi kesalahan sistem saat menghapus data.", "error")

    return redirect(request.referrer or url_for('index'))


# ─── Export Excel ──────────────────────────────────────────────────────────────

def _build_excel_workbook(transactions, user_label, month_label):
    """Buat workbook openpyxl dari list transaksi."""
    wb = openpyxl.Workbook()

    # ── Warna & Style ──────────────────────────────────────────────────
    header_fill  = PatternFill("solid", fgColor="1C1F2E")
    pemasukan_fill   = PatternFill("solid", fgColor="D1FAE5")   # hijau muda
    pengeluaran_fill = PatternFill("solid", fgColor="FEE2E2")   # merah muda
    investasi_fill   = PatternFill("solid", fgColor="DBEAFE")   # biru muda
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font(bold=True):
        return Font(name="Calibri", bold=bold, color="FFFFFF", size=10)

    def cell_font(bold=False, color="000000"):
        return Font(name="Calibri", bold=bold, color=color, size=10)

    # ── Sheet 1: Transaksi ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Transaksi"

    # Judul
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"Laporan Keuangan — {user_label} | {month_label}"
    title_cell.font  = Font(name="Calibri", bold=True, size=13, color="1C1F2E")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    gen_cell = ws["A2"]
    gen_cell.value = f"Dibuat pada: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    gen_cell.font  = Font(name="Calibri", italic=True, size=9, color="888888")
    gen_cell.alignment = Alignment(horizontal="center")

    # Header kolom
    headers = ["No", "ID Transaksi", "Tanggal & Waktu", "User", "Username",
               "Tipe", "Item / Keterangan", "Kategori", "Nominal (Rp)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.fill      = header_fill
        cell.font      = hdr_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[4].height = 22

    # Isi data
    tipe_fill_map = {
        "pemasukan":   pemasukan_fill,
        "pengeluaran": pengeluaran_fill,
        "investasi":   investasi_fill,
    }

    for row_no, tx in enumerate(transactions, 1):
        r = row_no + 4
        tipe_str = (tx.get("tipe") or "").lower()
        row_fill = tipe_fill_map.get(tipe_str, None)

        nama = tx.get("first_name") or ""
        if tx.get("last_name"):
            nama += " " + tx["last_name"]

        ts = tx.get("timestamp")
        ts_str = ts.strftime("%d/%m/%Y %H:%M") if ts else "-"

        nominal = tx.get("nominal") or 0

        row_data = [
            row_no,
            f"T-{tx.get('id', '')}",
            ts_str,
            nama.strip() or "-",
            f"@{tx.get('username')}" if tx.get("username") else "-",
            (tipe_str or "-").capitalize(),
            tx.get("item") or "-",
            (tx.get("kategori") or "-").replace("_", " ").capitalize(),
            nominal,
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.border = border
            cell.font   = cell_font()
            if row_fill:
                cell.fill = row_fill
            # Nominal rata kanan
            if col_idx == 9:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            elif col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto lebar kolom
    col_widths = [6, 14, 20, 20, 18, 15, 35, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"  # Freeze header

    # ── Sheet 2: Summary ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Ringkasan")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 22

    summary_title = ws2.cell(row=1, column=1, value="Ringkasan Keuangan")
    summary_title.font = Font(name="Calibri", bold=True, size=13, color="1C1F2E")
    ws2.merge_cells("A1:B1")
    summary_title.alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 24

    ws2.cell(row=2, column=1, value="Filter").font = Font(bold=True, size=9, color="888888")
    ws2.cell(row=2, column=2, value=f"{user_label} | {month_label}").font = Font(size=9, color="888888")

    total_pemasukan   = sum(t.get("nominal", 0) or 0 for t in transactions if (t.get("tipe") or "").lower() == "pemasukan")
    total_pengeluaran = sum(t.get("nominal", 0) or 0 for t in transactions if (t.get("tipe") or "").lower() == "pengeluaran")
    total_investasi   = sum(t.get("nominal", 0) or 0 for t in transactions if (t.get("tipe") or "").lower() == "investasi")
    saldo_cash   = total_pemasukan - total_pengeluaran
    saldo_bersih = total_pemasukan - total_pengeluaran - total_investasi

    summary_rows = [
        ("Total Pemasukan",   total_pemasukan,   "3D9970"),
        ("Total Pengeluaran", total_pengeluaran, "E74C3C"),
        ("Total Investasi",   total_investasi,   "3498DB"),
        ("Saldo Cash (Pem.-Peng.)", saldo_cash,  "2C3E50"),
        ("Saldo Bersih (inc. Investasi)", saldo_bersih, "8E44AD"),
    ]
    for ridx, (label, val, color) in enumerate(summary_rows, 4):
        lc = ws2.cell(row=ridx, column=1, value=label)
        lc.font   = Font(name="Calibri", bold=True, size=10, color=color)
        lc.border = border
        vc = ws2.cell(row=ridx, column=2, value=val)
        vc.font          = Font(name="Calibri", size=10)
        vc.border        = border
        vc.number_format = '#,##0'
        vc.alignment     = Alignment(horizontal="right")

    ws2.cell(row=10, column=1, value="Total Transaksi").font = Font(bold=True, size=10)
    ws2.cell(row=10, column=2, value=len(transactions)).alignment = Alignment(horizontal="center")
    ws2.cell(row=10, column=2).font = Font(size=10)

    return wb


@app.route('/export/excel')
def export_excel():
    """
    Export Excel transaksi.
    Query params opsional: month (YYYY-MM), user_id (angka).
    Contoh: /export/excel?month=2026-04&user_id=123456789
    """
    # Ambil filter dari query string (sama persis dengan filter utama)
    month_str, user_id_filter, uid = _parse_filters(request)

    # Tentukan label untuk nama file & judul sheet
    user_label  = "Semua User"
    month_label = month_str

    if uid:
        all_users = database.get_all_users()
        matched = next((u for u in all_users if u["user_id"] == uid), None)
        if matched:
            nama = matched["first_name"] or ""
            if matched.get("last_name"):
                nama += " " + matched["last_name"]
            user_label = nama.strip() or str(uid)

    try:
        transactions = database.get_all_transactions_export(month_str=month_str, user_id=uid)
    except Exception as e:
        print(f"Error in dashboard export: {e}")
        flash("Terjadi kesalahan sistem saat menyiapkan file export.", "error")
        return redirect(url_for('index'))

    wb = _build_excel_workbook(transactions, user_label, month_label)

    # Simpan ke buffer memori
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_user  = user_label.replace(" ", "_").replace("/", "-")
    filename   = f"transaksi_{safe_user}_{month_label}.xlsx"

    return send_file(
        buf,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == '__main__':
    app.run(debug=True, port=5000)
